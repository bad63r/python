import os
from PyLTSpice.LTSpiceBatch import SimCommander
import ltspice
import numpy as np
import openpyxl
import sys
from progress.bar import ChargingBar
from contextlib import contextmanager

#suppressing terminal output
@contextmanager
def suppress_stdout():
    with open(os.devnull, "w") as devnull:
        old_stdout = sys.stdout
        sys.stdout = devnull
        try:  
            yield
        finally:
            sys.stdout = old_stdout

# determine if running as script or exe file
if getattr(sys, 'frozen', False):
    meAbsPath = os.path.dirname(sys.executable)
elif __file__:
    meAbsPath = os.path.dirname(__file__)
#make a list of all excel files
files_in_directory = os.listdir(meAbsPath)
# creating sample source file path
filtered_files = []
filtered_files = [file for file in files_in_directory if file.endswith(".xlsx")]

if (len(filtered_files) > 1):
    print("ERROR: There must be only one excel file in directory from which you are running this python script!")
    sys.exit(1)
elif (len(filtered_files) == 0):
    print("ERROR: There is no excel file in the directory!")
    sys.exit(1)
excel_file_name = filtered_files[0]
sample_source_path = os.path.join(meAbsPath, excel_file_name)
# to open the workbook, workbook object is created
wb_obj = openpyxl.load_workbook(sample_source_path)
# get workbook active sheet object
for existing_sheet in wb_obj.sheetnames:
    sheet= wb_obj[existing_sheet]
    # get row numbers in excel read file
    row_rd = sheet.max_row
    # define empty list for every parameter
    Cs_param_list    = []
    Lpar_param_list  = []
    Cpar_param_list  = []
    Crail_param_list = []
    # extract data samples for Cs_param_list
    for i in range(2, row_rd + 1):
        cell_rd = sheet.cell(row = i, column = 2)
        Cs_param_list.append(str(cell_rd.value) + "pF")
    # extract data samples for Lpar_param_list
    for i in range(2, row_rd + 1):
        cell_rd = sheet.cell(row = i, column = 3)
        Lpar_param_list.append(str(cell_rd.value) + "uH")
    # extract data samples for Cpar_param_list
    for i in range(2, row_rd + 1):
        cell_rd = sheet.cell(row = i, column = 4)
        Cpar_param_list.append(str(cell_rd.value) + "pF")
    # extract data samples for Crail_param_list
    for i in range(2, row_rd + 1):
        cell_rd = sheet.cell(row = i, column = 5)
        Crail_param_list.append(str(cell_rd.value) + "pF")

    sim_cnt = 1
    # loading LTSpice file
    with suppress_stdout():
        LTS_file = SimCommander(meAbsPath + "\\Hamon_RLC_Transfer.asc")
    # setting progress bar
    bar = ChargingBar('Processing excel sheet ' + existing_sheet ,max=len(Cs_param_list))
    for i in range(0, len(Cs_param_list)):
        try:
            files_in_directory = os.listdir(meAbsPath)
            filtered_files = [file for file in files_in_directory if file.endswith(".raw")]
            for file in filtered_files:
                path_to_file = os.path.join(meAbsPath, file)
                os.remove(path_to_file)
            filtered_files = [file for file in files_in_directory if file.endswith(".net")]
            for file in filtered_files:
                path_to_file = os.path.join(meAbsPath, file)
                os.remove(path_to_file)
            filtered_files = [file for file in files_in_directory if file.endswith(".log")]
            for file in filtered_files:
                path_to_file = os.path.join(meAbsPath, file)
                os.remove(path_to_file)
            filtered_files = [file for file in files_in_directory if file.endswith(".fail")]
            for file in filtered_files:
                path_to_file = os.path.join(meAbsPath, file)
                os.remove(path_to_file)
        except:
            "Some files can not be deleted as they are occupied with other program."
        # changing parameters of the simulation
        LTS_file.set_parameters(Cs=Cs_param_list[i])
        LTS_file.set_parameters(Lpar=Lpar_param_list[i])
        LTS_file.set_parameters(Cpar=Cpar_param_list[i])
        LTS_file.set_parameters(Crail=Crail_param_list[i])
        # running LTSpice simulation and waiting to complete
        with suppress_stdout():
            LTS_file.run()
            LTS_file.wait_completion()
        # parsing data which was created out of simulation
        ltspice_raw_file = 'Hamon_RLC_Transfer_' + str(sim_cnt) + '.raw'
        sim_data = ltspice.Ltspice(ltspice_raw_file)
        sim_data.parse()
        # get frequency
        freq = sim_data.get_frequency()
        # get data needed to calculate module of transfer
        Vpar = sim_data.get_data('V(vpar)')
        Vpar = np.array(Vpar)
        Vser = sim_data.get_data('V(vser)')
        Vser = np.array(Vser)
        Irs0 = sim_data.get_data('I(Rs0)')
        Irs0 = np.array(Irs0)
        Irp0 = sim_data.get_data('I(Rp0)')
        Irp0 = np.array(Irp0)
        #(-V(vser)/I(Rs0))/(-V(vpar)/I(Rp0))
        data_transfera = (-Vser/Irs0)/(-Vpar/Irp0)
        moduo_transfera = np.abs(data_transfera)
        angle_transfer = np.angle(data_transfera,deg=True)

        cell = sheet.cell(row=1+sim_cnt, column=6)
        cell.value = str(moduo_transfera[len(data_transfera)-1])
        cell = sheet.cell(row=1+sim_cnt, column=7)
        cell.value = str(angle_transfer[len(angle_transfer)-1])
        sim_cnt = sim_cnt +1
        bar.next()
        if (sim_cnt == len(Cs_param_list) + 1):
            bar.finish()

        # print("data transfera: {}".format(data_transfera[len(data_transfera)-1]))
        # print("moduo transfera: {}".format(moduo_transfera[len(moduo_transfera)-1]))
        # print("ugao transfera: {}".format(angle_transfer[len(angle_transfer)-1]))
        wb_obj.save(excel_file_name)
print("Successfully calculated data!")
input("   | To exit, press any button...")

